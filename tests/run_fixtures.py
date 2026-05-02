#!/usr/bin/env python3
"""
Verifica o parser contra casos de edital organizados por pasta.

Formato novo:
  editais/<caso>/fixture.json
  editais/<caso>/<documento_principal>.pdf

Formato legado ainda suportado:
  tests/fixtures.json com "arquivo" relativo ao diretório raiz de editais.
"""
import argparse
import io
import json
import os
import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

try:
    import pdfplumber
except ImportError:
    print("ERRO: pdfplumber nao instalado. Execute: pip install pdfplumber")
    sys.exit(2)

try:
    import docx
except ImportError:
    docx = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

from parser_edital import analisar_sem_api

LEGACY_FIXTURES_PATH = ROOT / "tests" / "fixtures.json"
PESOS_CONFIANCA = {
    "numero_edital": 12,
    "orgao": 12,
    "cnpj": 8,
    "modalidade": 10,
    "objeto": 18,
    "valor": 12,
    "data_abertura": 12,
    "prazo_vigencia": 6,
    "criterio_julgamento": 6,
    "documentos_habilitacao": 4,
}


def _check(valor_real, regra):
    real_str = str(valor_real or "")
    if isinstance(regra, str):
        ok = real_str == regra
        return ok, f"'{real_str}' == '{regra}'"
    if isinstance(regra, dict):
        if "contem" in regra:
            termos = regra["contem"]
            ok = all(t.upper() in real_str.upper() for t in termos)
            return ok, f"'{real_str}' contem {termos}"
        if "min" in regra or "max" in regra:
            try:
                num = float(real_str.replace(".", "").replace(",", "."))
            except ValueError:
                return False, f"'{real_str}' nao e numerico"
            mn, mx = regra.get("min"), regra.get("max")
            if mn is not None and num < mn:
                return False, f"{num} < minimo {mn}"
            if mx is not None and num > mx:
                return False, f"{num} > maximo {mx}"
            return True, f"{num} em [{mn}, {mx}]"
    return False, f"regra desconhecida: {regra!r}"


def _check_nao(valor_real, proibidos):
    real_upper = str(valor_real or "").upper()
    for p in proibidos:
        if p.upper() in real_upper:
            return False, f"'{valor_real}' contem valor proibido '{p}'"
    return True, f"'{valor_real}' nao contem {proibidos}"


def _valor_igual_ou_contido(valor_real, regra):
    real_str = str(valor_real or "")
    if isinstance(regra, str):
        return real_str == regra
    if isinstance(regra, dict) and "contem" in regra:
        termos = regra["contem"]
        return all(t.upper() in real_str.upper() for t in termos)
    return False


def _peso_perdido(valor_real, regra, peso):
    return 0 if _valor_igual_ou_contido(valor_real, regra) else peso


def _resolver_pdf(caso_dir: Path, arquivo_principal: str) -> Path:
    caminho = caso_dir / arquivo_principal
    if caminho.exists():
        return caminho
    raise FileNotFoundError(f"arquivo principal nao encontrado: {caminho}")


def _ler_pdf_texto(pdf_path: Path, max_pgs):
    with open(pdf_path, "rb") as f:
        data = f.read()
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        paginas = pdf.pages[:max_pgs] if max_pgs else pdf.pages
        return "\n".join(p.extract_text() or "" for p in paginas)


def _texto_odt(conteudo: bytes) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(conteudo)) as zf:
            with zf.open("content.xml") as fh:
                root = ET.fromstring(fh.read())
    except Exception:
        return ""

    ns = {"text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0"}
    partes = []
    for tag in (".//text:h", ".//text:p"):
        for node in root.findall(tag, ns):
            trecho = "".join(node.itertext()).strip()
            if trecho:
                partes.append(trecho)
    return "\n".join(partes).strip()


def _texto_docx(conteudo: bytes) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(conteudo)) as zf:
            with zf.open("word/document.xml") as fh:
                root = ET.fromstring(fh.read())
    except Exception:
        return ""

    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    partes = []
    for node in root.findall(".//w:t", ns):
        trecho = (node.text or "").strip()
        if trecho:
            partes.append(trecho)
    return "\n".join(partes).strip()


def _extrair_texto_por_tipo(nome: str, conteudo: bytes) -> tuple[str, str]:
    nome_lower = nome.lower()
    if nome_lower.endswith(".pdf"):
        return "pdf", _ler_pdf_texto_bytes(conteudo)
    if nome_lower.endswith(".docx") and docx is not None:
        texto_xml = _texto_docx(conteudo)
        if texto_xml.strip():
            return "docx", texto_xml
        doc = docx.Document(io.BytesIO(conteudo))
        return "docx", "\n".join(p.text for p in doc.paragraphs)
    if nome_lower.endswith((".xlsx", ".xls")) and openpyxl is not None:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
        linhas = []
        for nome_aba in wb.sheetnames:
            ws = wb[nome_aba]
            linhas.append(f"[Aba: {nome_aba}]")
            for row in ws.iter_rows(values_only=True):
                linha = "\t".join("" if c is None else str(c) for c in row)
                if linha.strip():
                    linhas.append(linha)
        return "planilha", "\n".join(linhas)
    if nome_lower.endswith(".odt"):
        return "odt", _texto_odt(conteudo)
    if nome_lower.endswith((".txt", ".md", ".csv")):
        return "texto", conteudo.decode("utf-8", errors="replace")
    return "outro", conteudo.decode("utf-8", errors="replace")


def _ler_pdf_texto_bytes(conteudo: bytes, max_pgs=None) -> str:
    with pdfplumber.open(io.BytesIO(conteudo)) as pdf:
        paginas = pdf.pages[:max_pgs] if max_pgs else pdf.pages
        return "\n".join(p.extract_text() or "" for p in paginas)


def _listar_arquivos_caso(caso_dir: Path) -> list[Path]:
    return sorted(
        [p for p in caso_dir.iterdir() if p.is_file() and p.name.lower() != "fixture.json"],
        key=lambda p: p.name.lower(),
    )


def _selecionar_principal(arquivos: list[Path], manifest: dict | None = None) -> Path:
    if manifest and manifest.get("arquivo_principal"):
        alvo = Path(manifest["arquivo_principal"])
        for arq in arquivos:
            if arq.name == alvo.name:
                return arq
    prioridades = (".pdf", ".docx", ".odt", ".xlsx", ".xls", ".txt", ".md", ".csv")
    for ext in prioridades:
        for arq in arquivos:
            if arq.name.lower().endswith(ext):
                return arq
    raise FileNotFoundError("nenhum arquivo principal encontrado no caso")


def _normalizar_texto(texto: str) -> str:
    return " ".join((texto or "").split())


def _extrair_cabecalho_estruturado(texto: str) -> str:
    bruto = _normalizar_texto(texto[:20000])
    def first(pat: str) -> str:
        m = re.search(pat, bruto, re.IGNORECASE)
        return m.group(1).strip() if m else ""

    numero = first(r"((?:PREG[AÃ]O\s+ELETR[ÔO]NICO|COMPARA[CÇ][AÃ]O DE PRE[CÇ]OS|CONCORR[ÊE]NCIA|DISPENSA|SRP|PE|CP|PP|TP)[^\n]{0,80}?\d{1,5}/\d{4})")
    orgao = first(r"((?:TRIBUNAL REGIONAL DO TRABALHO[^.\n]{0,140}|CONSELHO REGIONAL DE PSICOLOGIA[^.\n]{0,140}|PREFEITURA MUNICIPAL[^.\n]{0,140}|INSTITUTO INTERAMERICANO[^.\n]{0,140}))")
    valor = first(r"(R\$\s*[\d\.\,]+)")
    data = first(r"(\d{2}/\d{2}/\d{4})")
    modalidade = first(r"((?:PREG[AÃ]O\s+ELETR[ÔO]NICO|COMPARA[CÇ][AÃ]O DE PRE[CÇ]OS|CONCORR[ÊE]NCIA\s+ELETR[ÔO]NICA|DISPENSA\s+ELETR[ÔO]NICA|PREG[AÃ]O))")
    criterio = first(r"((?:MENOR PRE[CÇ]O|MAIOR DESCONTO|MELHOR T[ÉE]CNICA|T[ÉE]CNICA E PRE[CÇ]O)[^\n]{0,40})")

    linhas = ["## FICHA DE LICITAÇÃO"]
    if numero:
        linhas.append(f"- Nº / Processo: {numero}")
    if orgao:
        linhas.append(f"- Órgão: {orgao}")
    if modalidade:
        linhas.append(f"- Modalidade: {modalidade}")
    if criterio:
        linhas.append(f"- Critério de Julgamento: {criterio}")
    if valor:
        linhas.append(f"- Valor Estimado Total: {valor}")
    if data:
        linhas.append(f"- Abertura das Propostas: {data}")
    if len(linhas) == 1:
        return ""
    return "\n".join(linhas) + "\n"


def _cargar_caso_legado(fx: dict, editais_dir: Path) -> dict:
    arquivo = fx["arquivo"]
    caso_dir = editais_dir
    pdf_path = caso_dir / arquivo
    if not pdf_path.exists():
        nome = Path(arquivo).name
        matches = list(editais_dir.rglob(nome))
        if len(matches) == 1:
            pdf_path = matches[0]
        elif len(matches) > 1:
            raise FileExistsError(
                f"arquivo ambiguo '{arquivo}' encontrado em mais de um local: "
                + ", ".join(str(p) for p in matches)
            )
        else:
            raise FileNotFoundError(f"arquivo nao encontrado: {pdf_path}")
    return {
        "id": fx.get("id"),
        "descricao": fx.get("descricao"),
        "caso_dir": str(pdf_path.parent),
        "arquivo_principal": pdf_path.name,
        "esperado": fx.get("esperado", {}),
        "nao_esperado": fx.get("nao_esperado", {}),
        "max_paginas": fx.get("max_paginas"),
        "nota": fx.get("nota"),
        "fonte": "legacy",
    }


def _cargar_caso_manifest(manifest_path: Path) -> dict:
    dados = json.loads(manifest_path.read_text(encoding="utf-8"))
    if "arquivo_principal" not in dados:
        arquivos = _listar_arquivos_caso(manifest_path.parent)
        principal = _selecionar_principal(arquivos, dados)
        dados["arquivo_principal"] = principal.name
    dados["caso_dir"] = str(manifest_path.parent)
    dados["fonte"] = "manifest"
    return dados


def _carregar_casos(editais_dir: Path) -> list[dict]:
    # Formato por pasta: cada subdiretório é um "caso".
    # Se existir `fixture.json`, usa como gabarito/manifest; se não existir,
    # cria um caso implícito (sem esperado) para ainda medir extração e confiança.
    if editais_dir.exists() and editais_dir.is_dir():
        subdirs = sorted([p for p in editais_dir.iterdir() if p.is_dir()], key=lambda p: p.name.lower())
        if subdirs:
            casos: list[dict] = []
            for d in subdirs:
                manifest = d / "fixture.json"
                if manifest.exists():
                    casos.append(_cargar_caso_manifest(manifest))
                    continue

                arquivos = _listar_arquivos_caso(d)
                if not arquivos:
                    raise FileNotFoundError(f"nenhum arquivo encontrado em {d}")
                principal = _selecionar_principal(arquivos, None)
                casos.append(
                    {
                        "id": d.name,
                        "descricao": f"Pasta {d.name} (sem fixture.json)",
                        "caso_dir": str(d),
                        "arquivo_principal": principal.name,
                        "esperado": {},
                        "nao_esperado": {},
                        "fonte": "implicit",
                    }
                )
            return casos

    dados = json.loads(LEGACY_FIXTURES_PATH.read_text(encoding="utf-8"))
    casos = []
    for fx in dados["fixtures"]:
        casos.append(_cargar_caso_legado(fx, editais_dir))
    return casos


def rodar_caso(caso: dict):
    ok_total = 0
    falha_total = 0
    registro = {
        "id": caso.get("id"),
        "descricao": caso.get("descricao"),
        "caso_dir": caso.get("caso_dir"),
        "arquivo_principal": caso.get("arquivo_principal"),
        "status": "ok",
        "extraido": None,
        "checks": [],
    }

    caso_dir = Path(caso["caso_dir"])
    arquivos = _listar_arquivos_caso(caso_dir)
    if not arquivos:
        raise FileNotFoundError(f"nenhum arquivo encontrado em {caso_dir}")
    pdf_path = _resolver_pdf(caso_dir, caso["arquivo_principal"])
    tamanho = pdf_path.stat().st_size
    registro["pdf_path"] = str(pdf_path)
    registro["tamanho_bytes"] = tamanho

    try:
        partes = []
        meta_arquivos = []
        texto_principal = ""
        for arq in arquivos:
            with open(arq, "rb") as f:
                conteudo = f.read()
            tipo, texto = _extrair_texto_por_tipo(arq.name, conteudo)
            texto = (texto or "").strip()
            meta_arquivos.append({"arquivo": arq.name, "tipo": tipo, "chars": len(texto)})
            if arq.name == pdf_path.name:
                texto_principal = texto
            if texto:
                partes.append(f"=== {arq.name} [{tipo}] ===\n{texto}")
        cabecalho = _extrair_cabecalho_estruturado(texto_principal)
        texto = (cabecalho + "\n\n" + "\n\n".join(partes)).strip()
        if caso.get("max_paginas") and pdf_path.suffix.lower() == ".pdf":
            with open(pdf_path, "rb") as f:
                texto_principal = _ler_pdf_texto_bytes(f.read(), caso.get("max_paginas"))
                cabecalho = _extrair_cabecalho_estruturado(texto_principal)
                texto = (cabecalho + "\n\n" + "\n\n".join(partes)).strip()
        resultado = analisar_sem_api(texto)
    except Exception as e:
        registro["status"] = "erro"
        registro["erro"] = str(e)
        raise

    registro["extraido"] = resultado
    registro["confianca_final"] = resultado.get("confianca")
    registro["score_final"] = resultado.get("score")
    registro["peso_total"] = sum(PESOS_CONFIANCA.get(campo, 0) for campo in caso.get("esperado", {}))
    registro["peso_perdido"] = 0
    registro["arquivos"] = meta_arquivos

    print(f"  Pasta: {caso_dir}")
    print(f"  Principal: {pdf_path.name}")

    for campo, regra in caso.get("esperado", {}).items():
        valor = resultado.get(campo, "N/A")
        ok, msg = _check(valor, regra)
        peso = PESOS_CONFIANCA.get(campo, 0)
        perdido = _peso_perdido(valor, regra, peso)
        registro["peso_perdido"] += perdido
        registro["checks"].append(
            {
                "tipo": "esperado",
                "campo": campo,
                "regra": regra,
                "peso": peso,
                "peso_perdido": perdido,
                "valor_real": valor,
                "ok": ok,
                "msg": msg,
            }
        )
        if ok:
            print(f"  OK {campo} [{peso}]: {msg}")
            ok_total += 1
        else:
            print(f"  XX {campo} [{peso}]: FALHOU - {msg}")
            print(f"     -> extraido: {valor!r}")
            print(f"     -> esperado: {regra!r}")
            falha_total += 1

    for campo, proibidos in caso.get("nao_esperado", {}).items():
        valor = resultado.get(campo, "")
        ok, msg = _check_nao(valor, proibidos)
        registro["checks"].append(
            {
                "tipo": "nao_esperado",
                "campo": campo,
                "proibidos": proibidos,
                "valor_real": valor,
                "ok": ok,
                "msg": msg,
            }
        )
        if ok:
            print(f"  OK {campo} (nao esperado): {msg}")
            ok_total += 1
        else:
            print(f"  XX {campo} (nao esperado): FALHOU - {msg}")
            falha_total += 1

    return ok_total, falha_total, registro


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--id")
    parser.add_argument("--report")
    parser.add_argument("--editais-dir")
    args = parser.parse_args()

    editais_dir = Path(args.editais_dir or os.getenv("EDITAIS_DIR", "C:/Users/lisia/Desktop/editais"))
    casos = _carregar_casos(editais_dir)

    if args.id:
        casos = [c for c in casos if c["id"] == args.id]
        if not casos:
            print(f"ERRO: caso '{args.id}' nao encontrado")
            sys.exit(2)

    print(f"Editais dir: {editais_dir}")
    print(f"Casos carregados: {len(casos)}")

    total_ok = 0
    total_falha = 0
    relatorio_casos = []

    for caso in casos:
        print(f"\n{'='*60}")
        print(f"[{caso['id']}] {caso['descricao']}")
        if caso.get("nota"):
            print(f"  Nota: {caso['nota']}")
        ok, falha, registro = rodar_caso(caso)
        total_ok += ok
        total_falha += falha
        relatorio_casos.append(registro)

    print(f"\n{'='*60}")
    print(f"Resultado: {total_ok}/{total_ok + total_falha} verificacoes OK")

    if args.report:
        relatorio = {
            "editais_dir": str(editais_dir),
            "pesos_confianca": PESOS_CONFIANCA,
            "total_ok": total_ok,
            "total_falha": total_falha,
            "casos": relatorio_casos,
        }
        Path(args.report).write_text(
            json.dumps(relatorio, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
        print(f"Relatorio gravado em: {args.report}")

    sys.exit(1 if total_falha else 0)


if __name__ == "__main__":
    main()
